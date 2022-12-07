from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import datetime

#################################################################
#
# test
#
def test():
    namn = "Alexander Baltatzis"
    epost = "alba@kth.se"
    timbelopp = 150
    kursledare = ('Daniel Bosk', 'dbosk@kth.se')
    chef = "Karl Meinke"
    tider = []
    tider.append({"datum":"2022-12-03",
                  'tid':"8-10",
                  'kurskod':'DD1321',
                  'typ':"handl",
                  'timmar':2,
                  'koeff':1.33,
                  'omr_tid':2*1.33,
                  'belopp':timbelopp*2*1.33})
    tider.append({"datum":"2022-12-04",
                  'tid':"8-10",
                  'kurskod':'DD1321',
                  'typ':"övning",
                  'timmar':2,
                  'koeff':3,
                  'omr_tid':2*3,
                  'belopp':timbelopp*2*3})
    tider.append({"datum":"2022-12-05",
                  'tid':"8-10",
                  'kurskod':'DD1321',
                  'typ':"handl",
                  'timmar':2.1,
                  'koeff':1.33,
                  'omr_tid':2.1*1.33,
                  'belopp':timbelopp*2.1*1.33})
    tider.append({"datum":"2022-12-05",
                  'tid':"10-12",
                  'kurskod':'DD1310',
                  'typ':"handl",
                  'timmar':2,
                  'koeff':1.33,
                  'omr_tid':2*1.33,
                  'belopp':timbelopp*2*1.33})

    make_excel(namn,      
               epost,     
               tider,
               kursledare,
               chef
               )
            
    

    

#################################################################
#
# make_excel
#
#  Anm: tider är en lista av dictionaries med fälten
#       datum, tid, kurskod, typ, timmar, koeef, omr_tid, belopp
#
def make_excel(namn, epost,# namn och epost på den timanställde
               tider,      # [{datum:text, tid:text, kurskod:text, typ:text, timmar:X, koeff:Y, omr_tid,X*Y, belopp:Z}, ...]
               kursledare, # (namn, epost)
               chef,       # Chefens namnförtydligande
               org = "JH",
               projekt = "1102",
               timbelopp = 150,
               filnamn = ""
               ):
    login = epost.replace("@kth.se", "")
    if filnamn == "" :
        filnamn = login + "_tid_" + datetime.date.today().strftime("%Y-%m-%d.xlsx")

    wb = Workbook()
    ark = wb.active


    #############################################################
    # Logo
    ark.title = login + " " + datetime.date.today().strftime("%Y-%b")
    logo = Image("kth.png")
    ark.add_image(logo, "A1")
    
    #############################################################
    # kolumnstorlekar
    ark.column_dimensions['A'].width = 16  # 'Schemalagd tid'
    ark.column_dimensions['B'].width = 9   # 'Typ'           
    ark.column_dimensions['C'].width = 7   # 'timmar'      
    ark.column_dimensions['D'].width = 8   # 'koeff'         
    ark.column_dimensions['E'].width = 15  # 'omräknad tid'        
    ark.column_dimensions['F'].width = 7   # 'Timlön'     
    ark.column_dimensions['G'].width = 9   # 'Belopp'        
    ark.column_dimensions['H'].width = 9

    #############################################################
    # Börja på rad 6
    rad = "6"
    ark['A' + rad] = "Timredovisning"
    ark['D' + rad] = "Namn"
    ark['E' + rad] = namn
    ark['E' + rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")

    rad = incr(rad)
    ark['D' + rad] = 'epost'
    ark['E' + rad] = epost
    ark['E' + rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")

    rad = incr(rad, 2)
    ark['A' + rad] = 'Kurskod'
    ark['B' + rad] = ''
    kurskoder = []
    for kol in tider:
        if 'kurskod' in kol and kol['kurskod'] not in kurskoder:
            ark['B' + rad].value += kol['kurskod'] + " "
            kurskoder.append( kol['kurskod'] )
    
    rad = incr(rad)
    ark['A' + rad] = 'Timmar ska anges inklusive förberedelsetid enligt schablon'
    rad = incr(rad)
    ark['A' + rad] = 'Ange typ av undervisning övning, handledning'

    rad = incr(rad, 2)
    ark['A' + rad] = 'Schemalagd tid'
    ark['B' + rad] = 'Typ'           
    ark['C' + rad] = 'Timmar'        
    ark['D' + rad] = 'koeff'         
    ark['E' + rad] = 'Omräknad tid'
    ark['F' + rad] = 'Timlön'        
    ark['G' + rad] = 'Belopp'
    
    for kol in ['C', 'D', 'E', 'F', 'G']:
        ark[kol+rad].alignment = Alignment(horizontal="right")
    

    #############################################################
    # Summering på sista raden 
    rad = incr(rad)   
    sist = incr(rad, len(tider))
    ark['E'+sist].font = Font(bold=True)  
    ark['G'+sist].font = Font(bold=True)  
    ark['E'+sist].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    #ark['G'+sist].fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

    #############################################################
    # Matris med timredovisningen
    for i, kol in enumerate(tider):
        ark['A'+rad].value = f"{kol['datum']} {kol['tid']:>5}"
        ark['B'+rad] = kol['typ']      
        ark['C'+rad] = kol['timmar']
        ark['D'+rad] = kol['koeff']    
        ark['E'+rad].value = round(kol['omr_tid'], 1)
        ark['F'+rad].value = timbelopp
        ark['G'+rad].value = round(kol['belopp'], 1)

        if i % 2 == 0:
            for kol in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ark[kol+rad].fill = PatternFill(start_color="00E0E0E0", end_color="00E0E0E0", fill_type="solid")
        
        if i == 0:
            tidsumma = "=ROUNDUP(SUM(E"+rad 
            ark['G'+sist].value = '=G'+rad 
        else:
            tidsumma += ',E'+rad 
            ark['G'+sist].value += '+G'+rad

        rad = incr(rad)

    ark['E'+sist].value = tidsumma + '),1)'

            
    #############################################################
    # Kontering
    rad = incr(sist, 3)
    ark['A'+rad].value = "Kontering"
    rad = incr(rad)
    ark['A'+rad].value = "Org.enhet"
    ark['A'+rad].fill  = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    ark['B'+rad].value = "Projekt"
    ark['B'+rad].fill = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    rad = incr(rad)
    ark['A'+rad].value = org
    ark['A'+rad].fill  = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")
    ark['B'+rad].value = projekt
    ark['B'+rad].fill  = PatternFill(start_color="00EEECE1", end_color="00EEECE1", fill_type="solid")

    #############################################################
    # Underskrift
    rad = incr(rad, 3)
    ark['A'+rad].value = "_______________________________________"
    ark['F'+rad].value = "Kursansvarig"
    rad = incr(rad)
    ark['A'+rad].value = "Ekonomisk attest " + chef
    ark['F'+rad].value = kursledare[0]
    rad = incr(rad)
    ark['F'+rad].value = kursledare[1]
    
    rad = incr(rad, 2)
    ark['A'+rad].value = "Underskriven blankett lämnas till HR"
    
    wb.save(filnamn)

#################################################################
#
# incr
#
def incr( rad, i=1):
    return str( int(rad) + i )
    
#################################################################
#
# main - test
#
if __name__ == "__main__":
    test()

